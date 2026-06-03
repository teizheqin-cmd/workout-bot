import os
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from datetime import datetime

# ── Path config ────────────────────────────────────────────────────────────────
# Set EXCEL_PATH in Railway environment variables to your OneDrive path,
# OR for Railway cloud (no OneDrive access), we save locally and sync manually.
# Default: saves in the same folder as the script.
EXCEL_PATH = os.environ.get("EXCEL_PATH", "workout_log.xlsx")

HEADERS = ["Date", "Workout Type", "Duration", "Exercises Summary",
           "Total Volume (kg)", "Intensity", "Feedback", "Original Message"]

# ── Create file if not exists ──────────────────────────────────────────────────
def _init_workbook():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Workout Log"

    # Header styling
    header_fill = PatternFill("solid", fgColor="2E4057")
    header_font = Font(bold=True, color="FFFFFF", size=11)

    for col, header in enumerate(HEADERS, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    # Column widths
    widths = [12, 18, 12, 45, 18, 12, 60, 50]
    for col, width in enumerate(widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = width

    ws.row_dimensions[1].height = 22
    wb.save(EXCEL_PATH)
    return wb

# ── Append one workout row ─────────────────────────────────────────────────────
def append_workout_row(result: dict, raw_input: str):
    try:
        wb = openpyxl.load_workbook(EXCEL_PATH)
    except FileNotFoundError:
        wb = _init_workbook()
        wb = openpyxl.load_workbook(EXCEL_PATH)

    ws = wb.active
    next_row = ws.max_row + 1

    # Alternate row color
    row_fill = PatternFill("solid", fgColor="F0F4F8") if next_row % 2 == 0 else PatternFill("solid", fgColor="FFFFFF")

    values = [
        result.get("date", ""),
        result.get("workout_type", ""),
        result.get("duration", ""),
        result.get("exercises_summary", ""),
        result.get("total_volume_kg", 0),
        result.get("intensity", ""),
        result.get("feedback", ""),
        raw_input
    ]

    for col, value in enumerate(values, 1):
        cell = ws.cell(row=next_row, column=col, value=value)
        cell.fill = row_fill
        cell.alignment = Alignment(wrap_text=True, vertical="top")

    ws.row_dimensions[next_row].height = 60
    wb.save(EXCEL_PATH)

# ── Get last 10 workouts for /summary ─────────────────────────────────────────
def get_summary() -> str:
    try:
        wb = openpyxl.load_workbook(EXCEL_PATH)
    except FileNotFoundError:
        return "📭 还没有任何运动记录！发送你的第一次运动吧 💪"

    ws = wb.active
    rows = list(ws.iter_rows(min_row=2, values_only=True))

    if not rows:
        return "📭 还没有任何运动记录！"

    last_10 = rows[-10:]
    lines = ["📊 *最近运动记录*\n"]

    for row in reversed(last_10):
        date_val = row[0] or "?"
        wtype    = row[1] or "?"
        duration = row[2] or "?"
        volume   = row[4] or 0
        intensity = row[5] or "?"
        lines.append(
            f"📅 `{date_val}` | {wtype} | {duration} | Vol: {volume}kg | {intensity}"
        )

    total = len(rows)
    lines.append(f"\n_共 {total} 次记录_")
    return "\n".join(lines)
